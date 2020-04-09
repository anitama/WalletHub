import openpyxl

# excel_path="D://InterviewPrep//testdata.xlsx"
# wb=openpyxl.load_workbook(excel_path)
# sheet=wb.get_sheet_by_name("Sheet1")
# rows=sheet.max_row
# print("total rows :" , rows)
# cols=sheet.max_column
# print("total cols :" , cols)
#
# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(row=r,column=c).value , end=" ")
#
#     print()

def total_rows(file_name,sheet_name):
    wb=openpyxl.load_workbook(file_name)
    sheet=wb.get_sheet_by_name(sheet_name)
    return sheet.max_row

def total_cols(path,sheet_name):
    wb=openpyxl.load_workbook(path)
    sheet=wb.get_sheet_by_name(sheet_name)
    return sheet.max_column

def read_data_from_excel(path,sheet_name,row_num, cols_num):
    wb=openpyxl.load_workbook(path)
    sheet=wb.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_num,column=cols_num).value

def write_data_into_excel(file_name,sheet_name,row_num,cols_num,result):
    wb=openpyxl.load_workbook(file_name)
    sheet=wb.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num,column=cols_num).value=result
    wb.save(file_name)
